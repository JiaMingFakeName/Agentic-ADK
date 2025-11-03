# Copyright (C) 2025 AIDC-AI
# This project incorporates components from the Open Source Software below.
# The original copyright notices and the licenses under which we received such components are set forth below for informational purposes.
#
# Open Source Software Licensed under the MIT License:
# --------------------------------------------------------------------
# 1. vscode-extension-updater-gitlab 3.0.1 https://www.npmjs.com/package/vscode-extension-updater-gitlab
# Copyright (c) Microsoft Corporation. All rights reserved.
# Copyright (c) 2015 David Owens II
# Copyright (c) Microsoft Corporation.
# Terms of the MIT:
# --------------------------------------------------------------------
# MIT License
#
# Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

from __future__ import annotations

from io import BytesIO
import logging
from pathlib import Path
from typing import Any, BinaryIO, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ali_agentic_adk_python.core.docloader.base import BaseLoader
from ali_agentic_adk_python.core.indexes import Document

LOGGER = logging.getLogger(__name__)


class ExcelDocLoader(BaseLoader):
    """Load Excel files (XLSX, XLS) and convert sheets into ``Document`` objects.
    
    This loader processes Excel workbooks and converts each sheet into a separate document.
    The content can be formatted as:
    - CSV-like text (default): Each row as comma-separated values
    - Markdown tables: Formatted as markdown table syntax
    - JSON: Each row as a JSON object
    
    Example:
        >>> loader = ExcelDocLoader("data.xlsx", output_format="markdown")
        >>> documents = loader.load()
        >>> for doc in documents:
        ...     print(f"Sheet: {doc.metadata['sheet_name']}")
        ...     print(doc.page_content)
    """

    def __init__(
        self, 
        file_path: Optional[str] = None,
        *,
        output_format: str = "csv",
        include_header: bool = True,
        max_rows: Optional[int] = None,
    ) -> None:
        """Initialize the Excel document loader.
        
        Args:
            file_path: Path to the Excel file
            output_format: Output format ('csv', 'markdown', or 'json')
            include_header: Whether to include the header row
            max_rows: Maximum number of rows to process per sheet (None for all)
        """
        self.file_path = file_path
        self.output_format = output_format.lower()
        self.include_header = include_header
        self.max_rows = max_rows
        
        if self.output_format not in ("csv", "markdown", "json"):
            raise ValueError(f"Invalid output_format: {output_format}. Must be 'csv', 'markdown', or 'json'")

    def load(self) -> list[Document]:
        """Load the Excel file and convert to documents.
        
        Returns:
            List of Document objects, one per sheet
            
        Raises:
            ValueError: If file_path is not set
            FileNotFoundError: If the file does not exist
        """
        if not self.file_path:
            raise ValueError("ExcelDocLoader requires `file_path` or metadata-driven loading.")
        return self._load_from_path(self.file_path)

    def fetch_content(self, document_meta: dict[str, Any]) -> list[Document]:
        """Load Excel file using metadata parameters.
        
        Args:
            document_meta: Dictionary containing:
                - file_path or filePath: Path to the file
                - input_stream, stream, or bytes: File content
                - metadata: Additional metadata to merge
                
        Returns:
            List of Document objects
        """
        metadata_hint = document_meta.get("metadata")

        file_path = document_meta.get("file_path") or document_meta.get("filePath")
        if file_path:
            documents = self._load_from_path(str(file_path))
        elif stream := (
            document_meta.get("input_stream")
            or document_meta.get("stream")
            or document_meta.get("binary_stream")
        ):
            documents = self._load_from_stream(stream)
        elif raw_bytes := document_meta.get("bytes"):
            documents = self._load_from_stream(BytesIO(raw_bytes))
        else:
            documents = super().fetch_content(document_meta)

        if metadata_hint:
            for doc in documents:
                doc.metadata.update(metadata_hint)
        return documents

    def _load_from_path(self, file_path: str) -> list[Document]:
        """Load Excel file from a file path.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of Document objects
            
        Raises:
            FileNotFoundError: If the file does not exist
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return self._extract_documents(workbook, str(path))

    def _load_from_stream(self, stream: BinaryIO | BytesIO) -> list[Document]:
        """Load Excel file from a byte stream.
        
        Args:
            stream: Binary stream containing Excel file data
            
        Returns:
            List of Document objects
        """
        workbook = openpyxl.load_workbook(stream, read_only=True, data_only=True)
        return self._extract_documents(workbook, None)

    def _extract_documents(
        self, 
        workbook: openpyxl.Workbook, 
        source: Optional[str]
    ) -> list[Document]:
        """Extract documents from an Excel workbook.
        
        Args:
            workbook: Opened Excel workbook
            source: Source file path (None if from stream)
            
        Returns:
            List of Document objects
        """
        documents: list[Document] = []
        
        for sheet_name in workbook.sheetnames:
            try:
                sheet = workbook[sheet_name]
                content = self._format_sheet(sheet)
                
                metadata = {
                    "source": source or "stream",
                    "sheet_name": sheet_name,
                    "format": "excel",
                    "output_format": self.output_format,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
                
                document = Document(page_content=content, metadata=metadata)
                documents.append(document)
                
            except Exception as exc:
                LOGGER.warning(f"Failed to process sheet '{sheet_name}': {exc}")
                continue
        
        return documents

    def _format_sheet(self, sheet: Worksheet) -> str:
        """Format a worksheet according to the output format.
        
        Args:
            sheet: Excel worksheet to format
            
        Returns:
            Formatted content as string
        """
        # Collect all rows
        rows = []
        row_count = 0
        
        for row in sheet.iter_rows(values_only=True):
            if self.max_rows and row_count >= self.max_rows:
                break
                
            # Skip empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
                
            # Convert cells to strings
            row_data = [str(cell) if cell is not None else "" for cell in row]
            rows.append(row_data)
            row_count += 1
        
        if not rows:
            return ""
        
        # Format based on output format
        if self.output_format == "csv":
            return self._format_as_csv(rows)
        elif self.output_format == "markdown":
            return self._format_as_markdown(rows)
        elif self.output_format == "json":
            return self._format_as_json(rows)
        else:
            return self._format_as_csv(rows)

    def _format_as_csv(self, rows: list[list[str]]) -> str:
        """Format rows as CSV text.
        
        Args:
            rows: List of rows, each row is a list of cell values
            
        Returns:
            CSV-formatted string
        """
        lines = []
        for row in rows:
            # Escape commas and quotes in cells
            escaped_row = []
            for cell in row:
                if "," in cell or '"' in cell:
                    escaped_cell = f'"{cell.replace(chr(34), chr(34) + chr(34))}"'
                    escaped_row.append(escaped_cell)
                else:
                    escaped_row.append(cell)
            lines.append(",".join(escaped_row))
        return "\n".join(lines)

    def _format_as_markdown(self, rows: list[list[str]]) -> str:
        """Format rows as a Markdown table.
        
        Args:
            rows: List of rows, each row is a list of cell values
            
        Returns:
            Markdown table string
        """
        if not rows:
            return ""
        
        lines = []
        
        # Header row
        if self.include_header and len(rows) > 0:
            header = rows[0]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("|" + "|".join(["---" for _ in header]) + "|")
            data_rows = rows[1:]
        else:
            # Generate column headers if not including first row as header
            num_cols = len(rows[0]) if rows else 0
            header = [f"Column {i+1}" for i in range(num_cols)]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("|" + "|".join(["---" for _ in header]) + "|")
            data_rows = rows
        
        # Data rows
        for row in data_rows:
            # Escape pipe characters in cells
            escaped_row = [cell.replace("|", chr(92) + "|") for cell in row]
            lines.append("| " + " | ".join(escaped_row) + " |")
        
        return "\n".join(lines)

    def _format_as_json(self, rows: list[list[str]]) -> str:
        """Format rows as JSON array of objects.
        
        Args:
            rows: List of rows, each row is a list of cell values
            
        Returns:
            JSON string
        """
        import json
        
        if not rows:
            return "[]"
        
        # Use first row as keys if include_header is True
        if self.include_header and len(rows) > 1:
            keys = rows[0]
            data_rows = rows[1:]
        else:
            # Generate keys
            num_cols = len(rows[0]) if rows else 0
            keys = [f"column_{i+1}" for i in range(num_cols)]
            data_rows = rows
        
        # Build array of objects
        result = []
        for row in data_rows:
            # Pad row if it's shorter than keys
            padded_row = row + [""] * (len(keys) - len(row))
            row_dict = {keys[i]: padded_row[i] for i in range(len(keys))}
            result.append(row_dict)
        
        return json.dumps(result, ensure_ascii=False, indent=2)


